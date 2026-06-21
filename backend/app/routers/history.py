"""修改历史：读取审计日志，展示对指标的全部变更（管理员直改 + 采纳/驳回建议）。"""
import io

from fastapi import APIRouter, Depends
from fastapi.responses import StreamingResponse
from sqlalchemy.orm import Session

from ..database import get_db
from ..security import require_admin
from ..models import AuditLog, User, Indicator
from ..schemas import HistoryEntry

router = APIRouter(tags=["修改历史"])

ACTION_LABELS = {
    "admin_create": "管理员新增", "admin_update": "管理员修改", "admin_delete": "管理员删除",
    "accept_add": "采纳·新增", "accept_edit": "采纳·修改", "accept_delete": "采纳·删除",
    "reject": "驳回建议",
}
FIELD_LABELS = {
    "identifier": "标识符", "name_cn": "中文名称", "name_en": "英文名称", "unit": "计量单位",
    "definition": "定义", "method": "计算方法", "description": "指标说明", "survey_method": "调查方法",
    "data_source": "数据来源", "frequency": "发布频率", "classification_id": "所属分类",
    "source_standard_id": "来源标准",
}


def _rows(db: Session, limit: int, indicator_id):
    q = db.query(AuditLog)
    if indicator_id:
        q = q.filter(AuditLog.entity_type == "indicator", AuditLog.entity_id == indicator_id)
    rows = q.order_by(AuditLog.id.desc()).limit(min(limit, 2000)).all()
    actor_ids = {r.actor_id for r in rows if r.actor_id}
    ind_ids = {r.entity_id for r in rows if r.entity_type == "indicator" and r.entity_id}
    actors = {u.id: u for u in db.query(User).filter(User.id.in_(actor_ids)).all()} if actor_ids else {}
    inds = {i.id: i for i in db.query(Indicator).filter(Indicator.id.in_(ind_ids)).all()} if ind_ids else {}
    out = []
    for r in rows:
        a = actors.get(r.actor_id)
        i = inds.get(r.entity_id) if r.entity_type == "indicator" else None
        detail = r.detail or {}
        out.append({
            "id": r.id, "created_at": r.created_at,
            "actor_name": a.display_name if a else None,
            "action": r.action, "entity_type": r.entity_type, "entity_id": r.entity_id,
            "indicator_name": (i.name_cn if i else detail.get("name_cn")),
            "detail": detail,
        })
    return out


def _summarize(detail):
    if not detail:
        return ""
    changes = detail.get("changes")
    if isinstance(changes, dict):
        parts = []
        for k, ov in changes.items():
            label = FIELD_LABELS.get(k, k)
            old = (ov.get("old") if isinstance(ov, dict) else "") or "（空）"
            new = (ov.get("new") if isinstance(ov, dict) else "") or "（空）"
            parts.append(f"{label}: {old} → {new}")
        return "；".join(parts)
    changed = detail.get("changed")
    if isinstance(changed, list):
        return "变更字段：" + "、".join(FIELD_LABELS.get(k, k) for k in changed)
    return ""


@router.get("/history", response_model=list[HistoryEntry], summary="修改历史（管理员）")
def list_history(limit: int = 300, indicator_id: int | None = None,
                 db: Session = Depends(get_db), admin: User = Depends(require_admin)):
    return [HistoryEntry(**r) for r in _rows(db, limit, indicator_id)]


@router.get("/history/export", summary="导出修改历史 Excel（管理员）")
def export_history(limit: int = 2000, indicator_id: int | None = None,
                   db: Session = Depends(get_db), admin: User = Depends(require_admin)):
    import openpyxl
    from openpyxl.styles import Font, Alignment

    rows = _rows(db, limit, indicator_id)
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "修改历史"
    ws.append(["时间", "操作类型", "指标", "操作人", "变更详情"])
    for c in ws[1]:
        c.font = Font(bold=True)
    for r in rows:
        ws.append([
            r["created_at"].strftime("%Y-%m-%d %H:%M:%S") if r["created_at"] else "",
            ACTION_LABELS.get(r["action"], r["action"]),
            r["indicator_name"] or (f"#{r['entity_id']}" if r["entity_id"] else ""),
            r["actor_name"] or "",
            _summarize(r["detail"]),
        ])
    for idx, w in enumerate([20, 12, 22, 12, 70], start=1):
        ws.column_dimensions[chr(64 + idx)].width = w
    for row in ws.iter_rows(min_row=2):
        row[4].alignment = Alignment(wrap_text=True, vertical="top")
    buf = io.BytesIO(); wb.save(buf); buf.seek(0)
    return StreamingResponse(
        buf, media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        headers={"Content-Disposition": "attachment; filename=modification_history.xlsx"})
